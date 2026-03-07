"""
Fetcher – downloads the Bank of England OIS yield-curve archive
and extracts SONIA OIS spot rates for the requested tenors.
"""

import io
import logging
import zipfile
from datetime import datetime, timedelta
from typing import Optional

import pandas as pd
import requests
from openpyxl import load_workbook

from .config import (
    ARCHIVE_FILES,
    BOE_OIS_ARCHIVE_URL,
    DATA_START_ROW,
    HEADER_ROW,
    HISTORY_YEARS,
    REQUEST_HEADERS,
    REQUEST_TIMEOUT,
    SPOT_CURVE_SHEET,
    TENOR_TO_COL,
    TENORS,
)

logger = logging.getLogger(__name__)


def _download_archive() -> zipfile.ZipFile:
    """Download the BoE OIS daily data archive and return an in-memory ZipFile."""
    logger.info("Downloading BoE OIS archive from %s …", BOE_OIS_ARCHIVE_URL)
    resp = requests.get(
        BOE_OIS_ARCHIVE_URL,
        headers=REQUEST_HEADERS,
        timeout=REQUEST_TIMEOUT,
    )
    resp.raise_for_status()
    logger.info(
        "Download complete – %s bytes (HTTP %s).",
        f"{len(resp.content):,}",
        resp.status_code,
    )
    return zipfile.ZipFile(io.BytesIO(resp.content))


def _parse_xlsx(
    zf: zipfile.ZipFile,
    filename: str,
    since_date: Optional[datetime] = None,
) -> pd.DataFrame:
    """
    Parse a single XLSX file from the archive, extracting the spot-curve sheet.

    Parameters
    ----------
    zf         : open ZipFile
    filename   : name of the XLSX inside the ZIP
    since_date : only keep rows on or after this date (optional)

    Returns
    -------
    pd.DataFrame  with columns [date, tenor_1y, tenor_2y, … tenor_7y]
    """
    logger.info("Reading %s / sheet '%s' …", filename, SPOT_CURVE_SHEET)
    wb = load_workbook(io.BytesIO(zf.read(filename)), read_only=True)
    ws = wb[SPOT_CURVE_SHEET]

    records = []
    for row_idx, row in enumerate(
        ws.iter_rows(min_row=DATA_START_ROW, values_only=False), start=DATA_START_ROW
    ):
        date_cell = row[0].value  # Column 1 (0-indexed → [0])
        if date_cell is None:
            continue

        # Normalise to datetime
        if isinstance(date_cell, str):
            date_val = datetime.strptime(date_cell, "%Y-%m-%d")
        elif isinstance(date_cell, datetime):
            date_val = date_cell
        else:
            continue

        if since_date and date_val < since_date:
            continue

        record = {"date": date_val}
        for tenor in TENORS:
            col_idx = TENOR_TO_COL[tenor] - 1  # 0-indexed
            val = row[col_idx].value if col_idx < len(row) else None
            record[f"tenor_{tenor}y"] = val

        records.append(record)

    wb.close()
    df = pd.DataFrame(records)
    logger.info("Parsed %d rows from %s.", len(df), filename)
    return df


def fetch_historical(years: int = HISTORY_YEARS) -> pd.DataFrame:
    """
    Download the full archive and return data for the last *years* years.

    Returns
    -------
    pd.DataFrame  [date, tenor_1y, … tenor_7y]
    """
    since = datetime.now() - timedelta(days=years * 365)
    logger.info("Fetching historical data since %s …", since.date())

    zf = _download_archive()
    frames = []
    for fname in ARCHIVE_FILES:
        if fname in zf.namelist():
            df = _parse_xlsx(zf, fname, since_date=since)
            if not df.empty:
                frames.append(df)
    zf.close()

    if not frames:
        logger.warning("No data extracted from archive.")
        return pd.DataFrame()

    combined = pd.concat(frames, ignore_index=True)
    combined.sort_values("date", inplace=True)
    combined.drop_duplicates(subset="date", keep="last", inplace=True)
    combined.reset_index(drop=True, inplace=True)
    logger.info(
        "Historical fetch complete: %d rows from %s to %s.",
        len(combined),
        combined["date"].min().date() if not combined.empty else "N/A",
        combined["date"].max().date() if not combined.empty else "N/A",
    )
    return combined


def fetch_latest(latest_db_date: Optional[str] = None) -> pd.DataFrame:
    """
    Download the archive and return only rows *after* latest_db_date.

    If latest_db_date is None, falls back to fetching the full history.
    This is the function called by the daily scheduler.

    Parameters
    ----------
    latest_db_date : str (YYYY-MM-DD) – the most recent date already in the DB

    Returns
    -------
    pd.DataFrame
    """
    if latest_db_date is None:
        return fetch_historical()

    since = datetime.strptime(latest_db_date, "%Y-%m-%d")
    logger.info("Incremental fetch for dates after %s …", latest_db_date)

    zf = _download_archive()
    # Only parse the most recent file for incremental updates
    fname = ARCHIVE_FILES[-1]
    df = pd.DataFrame()
    if fname in zf.namelist():
        df = _parse_xlsx(zf, fname, since_date=since)
        # Exclude the date we already have – only truly new rows
        if not df.empty:
            df = df[df["date"] > pd.Timestamp(latest_db_date)]
    zf.close()

    logger.info("Incremental fetch found %d new rows.", len(df))
    return df
